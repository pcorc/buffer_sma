"""
Excel export utilities for consolidated workbook output.
"""

import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def export_consolidated_workbook(results_list, summary_df, output_dir, run_name='backtest',
                                df_regimes=None, regime_df=None, capture_ratios=None,
                                trigger_summary=None, selection_summary=None, month_summary=None):
    """
    Export all backtest results to a single Excel workbook with multiple tabs.

    Workbook Structure:
    - Tab 1: Summary (all iterations with key metrics)
    - Tab 2: Trade Log (all trades with iteration details)
    - Tab 3: Regime Analysis (regime-specific performance)
    - Tab 4: Capture Ratios (upside/downside capture)
    - Tab 5: Trigger Summary (aggregated by trigger type)
    - Tab 6: Selection Summary (aggregated by selection algo)
    - Tab 7+: Daily Time Series for each iteration (with regime data)

    Parameters:
        results_list: List of result dicts from run_single_ticker_backtest
        summary_df: Consolidated summary DataFrame
        output_dir: Directory path for output file
        run_name: Name prefix for the workbook file
        df_regimes: Optional DataFrame with Date and Regime columns
        regime_df: Optional DataFrame with regime-specific analysis
        capture_ratios: Optional DataFrame with capture ratios
        trigger_summary: Optional DataFrame with trigger aggregation
        selection_summary: Optional DataFrame with selection aggregation

    Returns:
        Path to created workbook
    """
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f'{run_name}_consolidated_{timestamp}.xlsx'
    filepath = os.path.join(output_dir, filename)

    # Create Excel writer
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:

        # =====================================================================
        # TAB 1: SUMMARY
        # =====================================================================

        # Add iteration number for clarity
        summary_export = summary_df.copy()
        summary_export.insert(0, 'iteration', range(1, len(summary_export) + 1))

        # Format percentages and round numbers
        pct_cols = [col for col in summary_export.columns if 'return' in col or 'excess' in col or 'dd' in col or 'volatility' in col]
        for col in pct_cols:
            if col in summary_export.columns:
                summary_export[col] = summary_export[col].round(4)

        if 'strategy_sharpe' in summary_export.columns:
            summary_export['strategy_sharpe'] = summary_export['strategy_sharpe'].round(2)

        summary_export.to_excel(writer, sheet_name='Summary', index=False)

        # Format Summary sheet
        worksheet = writer.sheets['Summary']

        # Header formatting
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        worksheet.freeze_panes = 'A2'

        # =====================================================================
        # TAB 2: TRADE LOG
        # =====================================================================

        all_trades = []
        for idx, result in enumerate(results_list, 1):
            if result['trade_history'].empty:
                continue

            trades = result['trade_history'].copy()

            # Add iteration details
            trades.insert(0, 'iteration', idx)
            trades['launch_month'] = result['launch_month']
            trades['trigger_type'] = result['trigger_type']
            trades['trigger_params'] = str(result['trigger_params'])
            trades['selection_algo'] = result['selection_algo']

            # Add selection reason
            trades['Selection_Reason'] = trades.apply(
                lambda row: _get_selection_reason(result['selection_algo'], row['To_Fund']),
                axis=1
            )

            # Add regime if available
            if df_regimes is not None:
                trades = trades.merge(df_regimes[['Date', 'Regime']], on='Date', how='left')
                trades['Regime'] = trades['Regime'].fillna('unknown')

            all_trades.append(trades)

        if all_trades:
            combined_trades = pd.concat(all_trades, ignore_index=True)

            # Reorder columns for better readability
            base_cols = [
                'iteration', 'Date', 'launch_month', 'trigger_type', 'trigger_params',
                'selection_algo', 'From_Fund', 'To_Fund', 'Trigger_Reason',
                'Selection_Reason', 'NAV_at_Switch'
            ]

            if df_regimes is not None and 'Regime' in combined_trades.columns:
                base_cols.insert(2, 'Regime')

            combined_trades = combined_trades[base_cols]

            combined_trades.to_excel(writer, sheet_name='Trade Log', index=False)

            # Format Trade Log sheet
            worksheet = writer.sheets['Trade Log']

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            worksheet.freeze_panes = 'A2'
        else:
            print(f"    ⚠ No trades to export")

        # =====================================================================
        # TAB 3: ROLL DATE SNAPSHOTS
        # =====================================================================

        print("  Creating Roll Date Snapshots tab...")

        # Extract roll date snapshots from first result's daily performance
        # (all iterations use same funds, so roll dates are the same)
        if results_list:
            # Get all unique roll dates from the first iteration
            first_result = results_list[0]
            daily_df = first_result['daily_performance']

            # Group by Outcome_Period_ID and get first row of each period (the roll date)
            roll_snapshots = []

            for period_id in daily_df['Outcome_Period_ID'].dropna().unique():
                period_data = daily_df[daily_df['Outcome_Period_ID'] == period_id]
                if len(period_data) > 0:
                    snapshot = period_data.iloc[0].copy()

                    # Extract fund name from Current_Fund
                    fund_name = snapshot.get('Current_Fund', '')

                    roll_snapshots.append({
                        'Fund': fund_name,
                        'Outcome_Period_ID': period_id,
                        'Roll_Date': snapshot.get('Roll_Date'),
                        'Total_Outcome_Days': snapshot.get('Total_Outcome_Days'),
                        'Starting_Fund_NAV': snapshot.get('Starting_Fund_NAV'),
                        'Starting_Ref_Index': snapshot.get('Starting_Ref_Index'),
                        'Original_Cap': snapshot.get('Original_Cap'),
                        'Original_Buffer': snapshot.get('Original_Buffer'),
                        'Starting_Downside_Before_Buffer': snapshot.get('Starting_Downside_Before_Buffer'),
                        'Fund_Cap_Value': snapshot.get('Starting_Fund_NAV', 0) * (1 + snapshot.get('Original_Cap', 0)) if snapshot.get('Starting_Fund_NAV') and snapshot.get('Original_Cap') else None,
                        'Ref_Index_Cap_Value': snapshot.get('Starting_Ref_Index', 0) * (1 + snapshot.get('Original_Cap', 0)) if snapshot.get('Starting_Ref_Index') and snapshot.get('Original_Cap') else None
                    })

            if roll_snapshots:
                roll_snapshot_df = pd.DataFrame(roll_snapshots)

                # Sort by Roll_Date
                roll_snapshot_df = roll_snapshot_df.sort_values('Roll_Date')

                # Round numeric columns
                numeric_cols = roll_snapshot_df.select_dtypes(include=['float64']).columns
                roll_snapshot_df[numeric_cols] = roll_snapshot_df[numeric_cols].round(6)

                roll_snapshot_df.to_excel(writer, sheet_name='Roll Date Snapshots', index=False)

                # Format sheet
                worksheet = writer.sheets['Roll Date Snapshots']

                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

                worksheet.freeze_panes = 'A2'

                print(f"    ✓ Roll Date Snapshots: {len(roll_snapshot_df)} roll dates")


        # =====================================================================
        # TAB 4: REGIME ANALYSIS
        # =====================================================================

        if regime_df is not None and not regime_df.empty:
            regime_export = regime_df.copy()

            # Round numeric columns
            numeric_cols = regime_export.select_dtypes(include=['float64']).columns
            regime_export[numeric_cols] = regime_export[numeric_cols].round(4)

            regime_export.to_excel(writer, sheet_name='Regime Analysis', index=False)

            # Format sheet
            worksheet = writer.sheets['Regime Analysis']

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            worksheet.freeze_panes = 'A2'

        # =====================================================================
        # TAB 5: CAPTURE RATIOS
        # =====================================================================

        if capture_ratios is not None and not capture_ratios.empty:
            capture_export = capture_ratios.copy()

            # Round numeric columns
            numeric_cols = capture_export.select_dtypes(include=['float64']).columns
            capture_export[numeric_cols] = capture_export[numeric_cols].round(4)

            capture_export.to_excel(writer, sheet_name='Capture Ratios', index=False)

            # Format sheet
            worksheet = writer.sheets['Capture Ratios']

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            worksheet.freeze_panes = 'A2'

        # =====================================================================
        # TAB 6: TRIGGER SUMMARY
        # =====================================================================

        if trigger_summary is not None and not trigger_summary.empty:

            trigger_export = trigger_summary.copy()

            # Round numeric columns
            numeric_cols = trigger_export.select_dtypes(include=['float64']).columns
            trigger_export[numeric_cols] = trigger_export[numeric_cols].round(4)

            trigger_export.to_excel(writer, sheet_name='By Trigger', index=False)

            # Format sheet
            worksheet = writer.sheets['By Trigger']

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            worksheet.freeze_panes = 'A2'

        # =====================================================================
        # TAB 7: SELECTION SUMMARY
        # =====================================================================

        if selection_summary is not None and not selection_summary.empty:

            selection_export = selection_summary.copy()

            # Round numeric columns
            numeric_cols = selection_export.select_dtypes(include=['float64']).columns
            selection_export[numeric_cols] = selection_export[numeric_cols].round(4)

            selection_export.to_excel(writer, sheet_name='By Selection', index=False)

            # Format sheet
            worksheet = writer.sheets['By Selection']

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            worksheet.freeze_panes = 'A2'

        # =====================================================================
        # TAB 8+: DAILY TIME SERIES FOR EACH ITERATION
        # =====================================================================

        for idx, result in enumerate(results_list, 1):
            # Create tab name
            trigger_short = result['trigger_type'].replace('_', '').replace('rebalance', 'reb')[:12]
            selection_short = result['selection_algo'].replace('_', '').replace('select', 'sel')[:12]
            tab_name = f"{idx}_{result['launch_month']}_{trigger_short}_{selection_short}"

            # Excel tab names max 31 characters
            if len(tab_name) > 31:
                tab_name = f"{idx}_{result['launch_month'][:3]}_{trigger_short[:8]}_{selection_short[:8]}"

            # Prepare daily data
            daily = result['daily_performance'].copy()

            # Add iteration info
            daily.insert(0, 'iteration', idx)
            daily.insert(1, 'launch_month', result['launch_month'])
            daily.insert(2, 'trigger_type', result['trigger_type'])
            daily.insert(3, 'selection_algo', result['selection_algo'])

            # Add regime information if available
            if df_regimes is not None:
                daily = daily.merge(df_regimes[['Date', 'Regime']], on='Date', how='left')
                daily['Regime'] = daily['Regime'].fillna('unknown')
                # Move Regime column to position 4 (after selection_algo)
                cols = list(daily.columns)
                cols.insert(4, cols.pop(cols.index('Regime')))
                daily = daily[cols]

            # Add trade indicator
            daily['Trade_Occurred'] = False
            if not result['trade_history'].empty:
                trade_dates = set(result['trade_history']['Date'])
                daily['Trade_Occurred'] = daily['Date'].isin(trade_dates)

            # Add returns
            daily['Strategy_Daily_Return'] = daily['Strategy_NAV'].pct_change()
            daily['SPY_Daily_Return'] = daily['SPY_NAV'].pct_change()
            daily['BUFR_Daily_Return'] = daily['BUFR_NAV'].pct_change()
            daily['Hold_Daily_Return'] = daily['Hold_NAV'].pct_change()

            # Add relative performance (indexed to 100)
            daily['Strategy_vs_SPY'] = (daily['Strategy_NAV'] / daily['SPY_NAV']) * 100
            daily['Strategy_vs_BUFR'] = (daily['Strategy_NAV'] / daily['BUFR_NAV']) * 100
            daily['Strategy_vs_Hold'] = (daily['Strategy_NAV'] / daily['Hold_NAV']) * 100

            # Round for readability
            numeric_cols = daily.select_dtypes(include=['float64']).columns
            daily[numeric_cols] = daily[numeric_cols].round(6)

            # Reorder columns for better analysis flow
            base_cols = [
                'Date', 'iteration', 'launch_month', 'trigger_type', 'selection_algo'
            ]

            if 'Regime' in daily.columns:
                base_cols.append('Regime')

            # Add identification and snapshot columns
            snapshot_cols = [
                'Current_Fund', 'Outcome_Period_ID', 'Roll_Date',
                'Starting_Fund_NAV', 'Starting_Ref_Index', 'Original_Cap', 'Original_Buffer',
                'Total_Outcome_Days', 'Remaining_Outcome_Days'
            ]

            # Current values
            current_cols = [
                'Current_Fund_NAV', 'Current_Ref_Index', 'Current_Remaining_Cap_Pct'
            ]

            # Calculated returns
            return_cols = [
                'Fund_Return_From_Roll', 'Ref_Index_Return_From_Roll'
            ]

            # NAV performance
            nav_cols = [
                'Strategy_NAV', 'SPY_NAV', 'BUFR_NAV', 'Hold_NAV'
            ]

            # Daily returns
            daily_return_cols = [
                'Strategy_Daily_Return', 'SPY_Daily_Return', 'BUFR_Daily_Return', 'Hold_Daily_Return'
            ]

            # Relative performance
            relative_cols = [
                'Strategy_vs_SPY', 'Strategy_vs_BUFR', 'Strategy_vs_Hold'
            ]

            # Cap and buffer metrics
            metrics_cols = [
                'Cap_Utilization', 'Cap_Remaining_Pct',
                'Downside_Before_Buffer_Pct', 'Starting_Downside_Before_Buffer'
            ]

            # Trading
            trade_cols = ['Trade_Occurred']

            # Combine all columns in logical order
            all_cols = base_cols + snapshot_cols + current_cols + return_cols + nav_cols + daily_return_cols + relative_cols + metrics_cols + trade_cols

            # Only include columns that exist
            ordered_cols = [col for col in all_cols if col in daily.columns]
            daily = daily[ordered_cols]

            # Write to Excel
            daily.to_excel(writer, sheet_name=tab_name, index=False)

            # Format sheet
            worksheet = writer.sheets[tab_name]

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Auto-adjust columns (but faster than full iteration)
            for column_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                worksheet.column_dimensions[column_letter].width = 15

            worksheet.freeze_panes = 'A2'

    # Count total tabs
    total_tabs = 2  # Summary + Trade Log
    total_tabs += 1  # Roll Date Snapshots (always included if results exist)
    if regime_df is not None and not regime_df.empty:
        total_tabs += 1
    if capture_ratios is not None and not capture_ratios.empty:
        total_tabs += 1
    if trigger_summary is not None and not trigger_summary.empty:
        total_tabs += 1
    if selection_summary is not None and not selection_summary.empty:
        total_tabs += 1
    if month_summary is not None and not month_summary.empty:
        total_tabs += 1
    total_tabs += len(results_list)

    return filepath

def _get_selection_reason(selection_algo: str, selected_fund: str) -> str:
    """
    Generate human-readable selection reason based on algorithm.

    Parameters:
        selection_algo: Name of selection function
        selected_fund: The fund that was selected

    Returns:
        Human-readable reason string
    """
    reasons = {
        'select_most_recent_launch': f'Most recent roll date ({selected_fund})',
        'select_remaining_cap': f'Highest remaining cap ({selected_fund})',
        'select_cap_utilization': f'Lowest cap utilization ({selected_fund})',
        'select_highest_outcome_and_cap': f'Best outcome days + cap score ({selected_fund})',
        'select_cost_analysis': f'Lowest cost per day of protection ({selected_fund})'
    }

    return reasons.get(selection_algo, f'{selection_algo} selected {selected_fund}')


def export_consolidated_workbook_append(results_list, summary_df, output_dir, run_name='backtest'):
    """
    Export all backtest results to a single Excel workbook with multiple tabs.

    Workbook Structure:
    - Tab 1: Summary (all iterations with key metrics)
    - Tab 2: Trade Log (all trades with iteration details)
    - Tab 3+: Daily Time Series for each iteration (with regime data if provided)

    Parameters:
        results_list: List of result dicts from run_single_ticker_backtest
        summary_df: Consolidated summary DataFrame
        output_dir: Directory path for output file
        run_name: Name prefix for the workbook file
        df_regimes: Optional DataFrame with Date and Regime columns

    Returns:
        Path to created workbook
    """
    # First create the timestamped version
    timestamped_path = export_consolidated_workbook(results_list, summary_df, output_dir, run_name)

    # Then create/update cumulative version
    cumulative_filename = f'{run_name}_cumulative.xlsx'
    cumulative_path = os.path.join(output_dir, cumulative_filename)

    if os.path.exists(cumulative_path):

        existing_summary = pd.read_excel(cumulative_path, sheet_name='Summary')

        new_summary = summary_df.copy()
        new_summary.insert(0, 'iteration', range(len(existing_summary) + 1, len(existing_summary) + len(new_summary) + 1))
        new_summary['run_timestamp'] = datetime.now().strftime("%Y%m%d_%H%M%S")

        combined_summary = pd.concat([existing_summary, new_summary], ignore_index=True)

        # Read existing trade log if it exists
        try:
            existing_trades = pd.read_excel(cumulative_path, sheet_name='Trade Log')
        except:
            existing_trades = pd.DataFrame()

        # Create combined trades
        all_trades = []
        start_iteration = len(existing_summary) + 1

        for idx, result in enumerate(results_list, start_iteration):
            if result['trade_history'].empty:
                continue

            trades = result['trade_history'].copy()
            trades.insert(0, 'iteration', idx)
            trades['launch_month'] = result['launch_month']
            trades['trigger_type'] = result['trigger_type']
            trades['selection_algo'] = result['selection_algo']
            trades['Selection_Reason'] = trades.apply(
                lambda row: _get_selection_reason(result['selection_algo'], row['To_Fund']),
                axis=1
            )
            all_trades.append(trades)

        if all_trades:
            new_trades = pd.concat(all_trades, ignore_index=True)
            if not existing_trades.empty:
                combined_trades = pd.concat([existing_trades, new_trades], ignore_index=True)
            else:
                combined_trades = new_trades
        else:
            combined_trades = existing_trades


        # Write updated cumulative workbook
        with pd.ExcelWriter(cumulative_path, engine='openpyxl') as writer:
            combined_summary.to_excel(writer, sheet_name='Summary', index=False)
            if not combined_trades.empty:
                combined_trades.to_excel(writer, sheet_name='Trade Log', index=False)

    else:
        # Just copy the timestamped version as the first cumulative
        import shutil
        shutil.copy(timestamped_path, cumulative_path)

    return cumulative_path


"""
Excel export utilities for consolidated workbook output with forward regime analysis.
"""

def export_main_consolidated_workbook(
        results_list, summary_df, output_dir, run_name='mainpy_consolidated',
        df_regimes=None, regime_df=None, capture_ratios=None,
        trigger_summary=None, selection_summary=None, month_summary=None,
        # Forward regime parameters
        df_forward_regimes=None, future_regime_df=None,
        optimal_3m=None, optimal_6m=None,
        intent_vs_regime_3m=None, intent_vs_regime_6m=None,
        robust_strategies_3m=None, robust_strategies_6m=None,
        ranked_6m_vs_spy=None, ranked_6m_vs_bufr=None
):
    """
    Export comprehensive backtest results to single Excel workbook.

    Enhanced with forward regime analysis tabs for both 3M and 6M horizons.
    """
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f'{run_name}_{timestamp}.xlsx'
    filepath = os.path.join(output_dir, filename)

    print(f"\n{'=' * 80}")
    print(f"EXPORTING CONSOLIDATED WORKBOOK")

    # Header formatting
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    tab_count = 0

    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:

        # =====================================================================
        # TAB 1: SUMMARY
        # =====================================================================
        summary_export = summary_df.copy()
        summary_export.insert(0, 'iteration', range(1, len(summary_export) + 1))

        # Round numeric columns
        pct_cols = [col for col in summary_export.columns if 'return' in col or 'excess' in col or 'dd' in col or 'volatility' in col]
        for col in pct_cols:
            if col in summary_export.columns:
                summary_export[col] = summary_export[col].round(4)
        if 'strategy_sharpe' in summary_export.columns:
            summary_export['strategy_sharpe'] = summary_export['strategy_sharpe'].round(2)

        summary_export.to_excel(writer, sheet_name='Summary', index=False)
        _format_sheet(writer.sheets['Summary'], header_fill, header_font)
        tab_count += 1

        # =====================================================================
        # TAB 2: FUTURE REGIME ANALYSIS (6M)
        # =====================================================================
        if future_regime_df is not None and not future_regime_df.empty:
            future_export = future_regime_df.copy()
            numeric_cols = future_export.select_dtypes(include=['float64']).columns
            future_export[numeric_cols] = future_export[numeric_cols].round(4)
            future_export.to_excel(writer, sheet_name='Future Regime Analysis', index=False)
            _format_sheet(writer.sheets['Future Regime Analysis'], header_fill, header_font)
            tab_count += 1

        # =====================================================================
        # TABS 3-5: OPTIMAL STRATEGIES BY REGIME (6M)
        # =====================================================================
        if optimal_6m:
            for regime in ['bull', 'bear', 'neutral']:
                if regime in optimal_6m and not optimal_6m[regime].empty:
                    regime_optimal = optimal_6m[regime].copy()
                    numeric_cols = regime_optimal.select_dtypes(include=['float64']).columns
                    regime_optimal[numeric_cols] = regime_optimal[numeric_cols].round(4)

                    sheet_name = f'Optimal-{regime.title()} (6M)'
                    regime_optimal.to_excel(writer, sheet_name=sheet_name, index=False)
                    _format_sheet(writer.sheets[sheet_name], header_fill, header_font)
                    tab_count += 1

        # =====================================================================
        # TABS 6-8: OPTIMAL STRATEGIES BY REGIME (3M)
        # =====================================================================
        if optimal_3m:
            for regime in ['bull', 'bear', 'neutral']:
                if regime in optimal_3m and not optimal_3m[regime].empty:
                    regime_optimal = optimal_3m[regime].copy()
                    numeric_cols = regime_optimal.select_dtypes(include=['float64']).columns
                    regime_optimal[numeric_cols] = regime_optimal[numeric_cols].round(4)

                    sheet_name = f'Optimal-{regime.title()} (3M)'
                    regime_optimal.to_excel(writer, sheet_name=sheet_name, index=False)
                    _format_sheet(writer.sheets[sheet_name], header_fill, header_font)
                    tab_count += 1

        # =====================================================================
        # TAB: INTENT VS FUTURE REGIME (6M)
        # =====================================================================
        if intent_vs_regime_6m is not None and not intent_vs_regime_6m.empty:
            intent_export = intent_vs_regime_6m.copy()
            numeric_cols = intent_export.select_dtypes(include=['float64']).columns
            intent_export[numeric_cols] = intent_export[numeric_cols].round(4)
            intent_export.to_excel(writer, sheet_name='Intent vs Future (6M)', index=False)
            _format_sheet(writer.sheets['Intent vs Future (6M)'], header_fill, header_font)
            tab_count += 1

        # =====================================================================
        # TAB: INTENT VS FUTURE REGIME (3M)
        # =====================================================================
        if intent_vs_regime_3m is not None and not intent_vs_regime_3m.empty:
            intent_export = intent_vs_regime_3m.copy()
            numeric_cols = intent_export.select_dtypes(include=['float64']).columns
            intent_export[numeric_cols] = intent_export[numeric_cols].round(4)
            intent_export.to_excel(writer, sheet_name='Intent vs Future (3M)', index=False)
            _format_sheet(writer.sheets['Intent vs Future (3M)'], header_fill, header_font)
            tab_count += 1

        # =====================================================================
        # TAB: ROBUST STRATEGIES (6M)
        # =====================================================================
        if robust_strategies_6m is not None and not robust_strategies_6m.empty:
            robust_export = robust_strategies_6m.copy()
            numeric_cols = robust_export.select_dtypes(include=['float64']).columns
            robust_export[numeric_cols] = robust_export[numeric_cols].round(4)
            robust_export.to_excel(writer, sheet_name='Robust Strategies (6M)', index=False)
            _format_sheet(writer.sheets['Robust Strategies (6M)'], header_fill, header_font)
            tab_count += 1

        # =====================================================================
        # TAB: ROBUST STRATEGIES (3M)
        # =====================================================================
        if robust_strategies_3m is not None and not robust_strategies_3m.empty:
            robust_export = robust_strategies_3m.copy()
            numeric_cols = robust_export.select_dtypes(include=['float64']).columns
            robust_export[numeric_cols] = robust_export[numeric_cols].round(4)
            robust_export.to_excel(writer, sheet_name='Robust Strategies (3M)', index=False)
            _format_sheet(writer.sheets['Robust Strategies (3M)'], header_fill, header_font)
            tab_count += 1

        # =====================================================================
        # TAB: RANKED BY VS BUFR (6M)
        # =====================================================================
        if ranked_6m_vs_bufr is not None and not ranked_6m_vs_bufr.empty:
            ranked_export = ranked_6m_vs_bufr.copy()
            numeric_cols = ranked_export.select_dtypes(include=['float64']).columns
            ranked_export[numeric_cols] = ranked_export[numeric_cols].round(4)
            ranked_export.to_excel(writer, sheet_name='Ranked vs BUFR (6M)', index=False)
            _format_sheet(writer.sheets['Ranked vs BUFR (6M)'], header_fill, header_font)
            tab_count += 1

        # =====================================================================
        # TAB: RANKED BY VS SPY (6M)
        # =====================================================================
        if ranked_6m_vs_spy is not None and not ranked_6m_vs_spy.empty:
            ranked_export = ranked_6m_vs_spy.copy()
            numeric_cols = ranked_export.select_dtypes(include=['float64']).columns
            ranked_export[numeric_cols] = ranked_export[numeric_cols].round(4)
            ranked_export.to_excel(writer, sheet_name='Ranked vs SPY (6M)', index=False)
            _format_sheet(writer.sheets['Ranked vs SPY (6M)'], header_fill, header_font)
            tab_count += 1

        # =====================================================================
        # ORIGINAL ANALYSIS TABS
        # =====================================================================

        if regime_df is not None and not regime_df.empty:
            regime_export = regime_df.copy()
            numeric_cols = regime_export.select_dtypes(include=['float64']).columns
            regime_export[numeric_cols] = regime_export[numeric_cols].round(4)
            regime_export.to_excel(writer, sheet_name='Current Regime Analysis', index=False)
            _format_sheet(writer.sheets['Current Regime Analysis'], header_fill, header_font)
            tab_count += 1

        if capture_ratios is not None and not capture_ratios.empty:
            capture_export = capture_ratios.copy()
            numeric_cols = capture_export.select_dtypes(include=['float64']).columns
            capture_export[numeric_cols] = capture_export[numeric_cols].round(4)
            capture_export.to_excel(writer, sheet_name='Capture Ratios', index=False)
            _format_sheet(writer.sheets['Capture Ratios'], header_fill, header_font)
            tab_count += 1

        if trigger_summary is not None and not trigger_summary.empty:
            trigger_export = trigger_summary.copy()
            numeric_cols = trigger_export.select_dtypes(include=['float64']).columns
            trigger_export[numeric_cols] = trigger_export[numeric_cols].round(4)
            trigger_export.to_excel(writer, sheet_name='By Trigger Type', index=False)
            _format_sheet(writer.sheets['By Trigger Type'], header_fill, header_font)
            tab_count += 1

        if selection_summary is not None and not selection_summary.empty:
            selection_export = selection_summary.copy()
            numeric_cols = selection_export.select_dtypes(include=['float64']).columns
            selection_export[numeric_cols] = selection_export[numeric_cols].round(4)
            selection_export.to_excel(writer, sheet_name='By Selection Algo', index=False)
            _format_sheet(writer.sheets['By Selection Algo'], header_fill, header_font)
            tab_count += 1

        if month_summary is not None and not month_summary.empty:
            month_export = month_summary.copy()
            numeric_cols = month_export.select_dtypes(include=['float64']).columns
            month_export[numeric_cols] = month_export[numeric_cols].round(4)
            month_export.to_excel(writer, sheet_name='By Launch Month', index=False)
            _format_sheet(writer.sheets['By Launch Month'], header_fill, header_font)
            tab_count += 1


    return filepath


def _format_sheet(worksheet, header_fill, header_font):
    """Apply consistent formatting to worksheet."""
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    worksheet.freeze_panes = 'A2'