"""
Excel workbook consolidation with comprehensive multi-tab outputs.

Creates a single Excel workbook containing:
- All simulation results
- Best strategies selected
- Results grouped by intent
- Regime analysis
- Summary statistics
"""

import pandas as pd
import os
from datetime import datetime
from typing import Dict, List


def create_comprehensive_workbook(
        summary_df: pd.DataFrame,
        best_strategies_df: pd.DataFrame,
        intent_groups: Dict[str, pd.DataFrame],
        regime_df: pd.DataFrame,
        intent_summary: pd.DataFrame,
        output_dir: str,
        timestamp: str = None
):
    """
    Create comprehensive Excel workbook with all results organized in tabs.

    Tabs created:
    1. All Results - Every simulation
    2. Best Strategies - The 4 selected top performers
    3. By Intent - Bullish - All bullish strategies
    4. By Intent - Bearish - All bearish strategies
    5. By Intent - Neutral - All neutral strategies
    6. By Intent - Cost Opt - All cost-optimized strategies
    7. Regime Analysis - Performance by market regime
    8. Summary Stats - High-level overview

    Parameters:
        summary_df: DataFrame with all simulation results
        best_strategies_df: DataFrame with best strategy per intent
        intent_groups: Dict of DataFrames grouped by intent
        regime_df: DataFrame with regime-specific analysis
        intent_summary: DataFrame with summary stats by intent
        output_dir: Directory to save the workbook
        timestamp: Optional timestamp string for filename
    """
    print("\n" + "=" * 80)
    print("CREATING COMPREHENSIVE EXCEL WORKBOOK")
    print("=" * 80)

    if timestamp is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    filename = f'backtest_results_comprehensive_{timestamp}.xlsx'
    filepath = os.path.join(output_dir, filename)

    # Create Excel writer
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:

        # =====================================================================
        # TAB 1: ALL RESULTS
        # =====================================================================
        print("\nWriting Tab 1: All Results...")

        # Select and order columns for display
        all_results_cols = [
            'strategy_intent',
            'launch_month',
            'trigger_type',
            'trigger_params',
            'selection_algo',
            'strategy_return',
            'strategy_ann_return',
            'strategy_sharpe',
            'strategy_volatility',
            'strategy_max_dd',
            'vs_bufr_excess',
            'vs_spy_excess',
            'vs_hold_excess',
            'num_trades',
            'start_date',
            'end_date'
        ]

        all_results = summary_df[all_results_cols].copy()
        all_results.to_excel(writer, sheet_name='All Results', index=False)

        # Format the worksheet
        worksheet = writer.sheets['All Results']
        _format_worksheet(worksheet, all_results)

        print(f"  ✓ All Results: {len(all_results)} rows")

        # =====================================================================
        # TAB 2: BEST STRATEGIES
        # =====================================================================
        print("\nWriting Tab 2: Best Strategies...")

        best_cols = [
            'strategy_intent',
            'launch_month',
            'trigger_type',
            'trigger_params',
            'selection_algo',
            'selection_criteria',
            'strategy_return',
            'strategy_ann_return',
            'strategy_sharpe',
            'strategy_volatility',
            'strategy_max_dd',
            'vs_bufr_excess',
            'vs_spy_excess',
            'num_trades',
            'start_date',
            'end_date'
        ]

        best_results = best_strategies_df[best_cols].copy()
        best_results.to_excel(writer, sheet_name='Best Strategies', index=False)

        worksheet = writer.sheets['Best Strategies']
        _format_worksheet(worksheet, best_results, highlight_best=True)

        print(f"  ✓ Best Strategies: {len(best_results)} rows")

        # =====================================================================
        # TABS 3-6: BY INTENT
        # =====================================================================
        intent_tab_names = {
            'bullish': 'By Intent - Bullish',
            'bearish': 'By Intent - Bearish',
            'neutral': 'By Intent - Neutral',
            'cost_optimized': 'By Intent - Cost Opt'
        }

        for intent, tab_name in intent_tab_names.items():
            if intent in intent_groups and not intent_groups[intent].empty:
                print(f"\nWriting Tab: {tab_name}...")

                intent_df = intent_groups[intent][all_results_cols].copy()
                intent_df.to_excel(writer, sheet_name=tab_name, index=False)

                worksheet = writer.sheets[tab_name]
                _format_worksheet(worksheet, intent_df)

                print(f"  ✓ {tab_name}: {len(intent_df)} rows")

        # =====================================================================
        # TAB 7: REGIME ANALYSIS
        # =====================================================================
        if not regime_df.empty:
            print("\nWriting Tab: Regime Analysis...")

            regime_cols = [
                'launch_month',
                'trigger_type',
                'selection_algo',
                'regime',
                'days_in_regime',
                'strategy_return',
                'spy_return',
                'bufr_return',
                'vs_bufr_excess',
                'num_trades'
            ]

            regime_results = regime_df[regime_cols].copy()
            regime_results.to_excel(writer, sheet_name='Regime Analysis', index=False)

            worksheet = writer.sheets['Regime Analysis']
            _format_worksheet(worksheet, regime_results)

            print(f"  ✓ Regime Analysis: {len(regime_results)} rows")

        # =====================================================================
        # TAB 8: SUMMARY STATS
        # =====================================================================
        print("\nWriting Tab: Summary Stats...")

        # Create summary statistics table
        summary_stats = _create_summary_stats_table(
            summary_df, best_strategies_df, intent_summary
        )

        summary_stats.to_excel(writer, sheet_name='Summary Stats', index=False)

        worksheet = writer.sheets['Summary Stats']
        _format_worksheet(worksheet, summary_stats, summary_table=True)

        print(f"  ✓ Summary Stats: {len(summary_stats)} rows")

    print("\n" + "=" * 80)
    print(f"✅ WORKBOOK SAVED: {filename}")
    print(f"   Location: {output_dir}")
    print(f"   Total tabs: 8")
    print("=" * 80 + "\n")

    return filepath


def _format_worksheet(worksheet, df: pd.DataFrame, highlight_best: bool = False, summary_table: bool = False):
    """
    Apply formatting to Excel worksheet.

    Parameters:
        worksheet: openpyxl worksheet object
        df: DataFrame that was written to worksheet
        highlight_best: If True, highlight best values in key columns
        summary_table: If True, apply special formatting for summary tables
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Header formatting
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Auto-adjust column widths
    for idx, col in enumerate(df.columns, 1):
        column_letter = get_column_letter(idx)

        # Calculate max width
        max_length = len(str(col))
        for cell in worksheet[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # Freeze top row
    worksheet.freeze_panes = 'A2'

    # Number formatting for specific columns
    for idx, col in enumerate(df.columns, 1):
        column_letter = get_column_letter(idx)

        if 'return' in col.lower() or 'excess' in col.lower() or 'volatility' in col.lower():
            # Percentage format
            for row in range(2, len(df) + 2):
                cell = worksheet[f'{column_letter}{row}']
                cell.number_format = '0.00%'

        elif 'sharpe' in col.lower():
            # Two decimal places
            for row in range(2, len(df) + 2):
                cell = worksheet[f'{column_letter}{row}']
                cell.number_format = '0.00'

        elif 'max_dd' in col.lower():
            # Percentage format (negative)
            for row in range(2, len(df) + 2):
                cell = worksheet[f'{column_letter}{row}']
                cell.number_format = '0.00%'

        elif 'num_trades' in col.lower():
            # Integer format
            for row in range(2, len(df) + 2):
                cell = worksheet[f'{column_letter}{row}']
                cell.number_format = '0'

    # Conditional formatting for returns (green positive, red negative)
    if highlight_best and not summary_table:
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        for idx, col in enumerate(df.columns, 1):
            if 'return' in col.lower() or 'excess' in col.lower():
                column_letter = get_column_letter(idx)

                for row in range(2, len(df) + 2):
                    cell = worksheet[f'{column_letter}{row}']
                    try:
                        value = float(cell.value) if cell.value is not None else 0
                        if value > 0:
                            cell.fill = green_fill
                        elif value < 0:
                            cell.fill = red_fill
                    except:
                        pass

    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in worksheet.iter_rows(min_row=1, max_row=len(df) + 1,
                                   min_col=1, max_col=len(df.columns)):
        for cell in row:
            cell.border = thin_border


def _create_summary_stats_table(
        summary_df: pd.DataFrame,
        best_strategies_df: pd.DataFrame,
        intent_summary: pd.DataFrame
) -> pd.DataFrame:
    """
    Create comprehensive summary statistics table.

    Parameters:
        summary_df: All results DataFrame
        best_strategies_df: Best strategies DataFrame
        intent_summary: Intent-level summary DataFrame

    Returns:
        DataFrame with summary statistics
    """
    summary_records = []

    # Overall statistics
    summary_records.append({
        'Category': 'OVERALL',
        'Metric': 'Total Simulations',
        'Value': len(summary_df),
        'Notes': f"{summary_df['strategy_intent'].nunique()} intent categories"
    })

    summary_records.append({
        'Category': 'OVERALL',
        'Metric': 'Avg Return',
        'Value': summary_df['strategy_return'].mean(),
        'Notes': f"Median: {summary_df['strategy_return'].median():.2%}"
    })

    summary_records.append({
        'Category': 'OVERALL',
        'Metric': 'Avg Sharpe',
        'Value': summary_df['strategy_sharpe'].mean(),
        'Notes': f"Median: {summary_df['strategy_sharpe'].median():.2f}"
    })

    summary_records.append({
        'Category': 'OVERALL',
        'Metric': '% Beat BUFR',
        'Value': (summary_df['vs_bufr_excess'] > 0).sum() / len(summary_df),
        'Notes': f"{(summary_df['vs_bufr_excess'] > 0).sum()} strategies"
    })

    summary_records.append({
        'Category': 'OVERALL',
        'Metric': '% Beat SPY',
        'Value': (summary_df['vs_spy_excess'] > 0).sum() / len(summary_df),
        'Notes': f"{(summary_df['vs_spy_excess'] > 0).sum()} strategies"
    })

    # Best strategies summary
    for _, row in best_strategies_df.iterrows():
        intent = row['strategy_intent'].upper()

        summary_records.append({
            'Category': f'BEST - {intent}',
            'Metric': 'Launch Month',
            'Value': row['launch_month'],
            'Notes': f"{row['trigger_type'][:30]}"
        })

        summary_records.append({
            'Category': f'BEST - {intent}',
            'Metric': 'Return',
            'Value': row['strategy_return'],
            'Notes': f"Ann: {row['strategy_ann_return']:.2%}"
        })

        summary_records.append({
            'Category': f'BEST - {intent}',
            'Metric': 'Sharpe',
            'Value': row['strategy_sharpe'],
            'Notes': f"MaxDD: {row['strategy_max_dd']:.2%}"
        })

        summary_records.append({
            'Category': f'BEST - {intent}',
            'Metric': 'vs BUFR',
            'Value': row['vs_bufr_excess'],
            'Notes': f"Trades: {int(row['num_trades'])}"
        })

    # Intent-level statistics
    if not intent_summary.empty:
        for _, row in intent_summary.iterrows():
            intent = row['intent'].upper()

            summary_records.append({
                'Category': f'INTENT - {intent}',
                'Metric': 'Count',
                'Value': int(row['count']),
                'Notes': f"{row['pct_beat_bufr']:.1f}% beat BUFR"
            })

            summary_records.append({
                'Category': f'INTENT - {intent}',
                'Metric': 'Avg Return',
                'Value': row['avg_return'],
                'Notes': f"Best: {row['best_return']:.2%}"
            })

            summary_records.append({
                'Category': f'INTENT - {intent}',
                'Metric': 'Avg Sharpe',
                'Value': row['avg_sharpe'],
                'Notes': f"Best: {row['best_sharpe']:.2f}"
            })

    summary_table = pd.DataFrame(summary_records)

    return summary_table


def export_results_csv(
        summary_df: pd.DataFrame,
        regime_df: pd.DataFrame,
        output_dir: str
):
    """
    Export results as separate CSV files (legacy support).

    Parameters:
        summary_df: Summary DataFrame
        regime_df: Regime analysis DataFrame
        output_dir: Output directory
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Export summary
    summary_filename = f'backtest_results_{timestamp}.csv'
    summary_path = os.path.join(output_dir, summary_filename)
    summary_df.to_csv(summary_path, index=False)

    # Export regime
    regime_filename = f'regime_analysis_{timestamp}.csv'
    regime_path = os.path.join(output_dir, regime_filename)
    regime_df.to_csv(regime_path, index=False)

    # Also save as "latest"
    summary_df.to_csv(os.path.join(output_dir, 'backtest_results_latest.csv'), index=False)
    regime_df.to_csv(os.path.join(output_dir, 'regime_analysis_latest.csv'), index=False)

    print(f"✓ CSV exports complete: {summary_filename}, {regime_filename}")